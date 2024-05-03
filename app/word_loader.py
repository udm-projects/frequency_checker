from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
import pickle
import os

from app.logger import logger
from config import path_to_serialization, path_to_xlsx_data_source
from app.word import Word, Words


def load_words(word_storage: Words):
    """Загрузка слов из файла и наполнение word_storage"""

    logger.info('Начало загрузки справочника слов из файла')
    if os.path.isfile(path_to_serialization):
        # быстрая загрузка
        load_words_from_pickle(word_storage)
    elif os.path.isfile(path_to_xlsx_data_source):
        for word_params in load_word_params_from_xlsx():
            word = Word(**word_params)
            word_storage.add_word(word)
        logger.info(f"Загружено {word_storage.get_count()} слов из xlsx файла")
        serialize_words_to_pickle(word_storage)
    else:
        raise NoWordFile("Отсутсвует файл со словами")


class NoWordFile(Exception):
    pass


def load_words_from_pickle(word_storage: Words):
    """Загрузка слов из файла pickle"""
    logger.debug("Начало загрузки слов из pickle")
    with open(path_to_serialization, 'rb') as f:
        word_storage.update_words(pickle.load(f))
    logger.info(f"Загружено {word_storage.get_count()} слов из pickle")


def load_word_params_from_xlsx() -> dict:
    """Загрузка слов из файла xlsx,
    возвращает генератор для создания объектов"""
    logger.debug("Начало загрузки слов из xlsx файла")
    word_sheet = get_word_sheet()
    return transform_sheet_data_to_word_dict(word_sheet)


def get_word_sheet() -> Worksheet:
    """Загрузка рабочего листа xlsx со словами"""
    wb = load_workbook(filename=path_to_xlsx_data_source)
    return wb['Sheet1']


def transform_sheet_data_to_word_dict(words: Worksheet) -> dict:
    """Преобразование рабочего листа xlsx со словами в словарь (генератор)"""
    for row in words.iter_rows(min_row=2):
        related_word_forms_list = (row[2].value.strip()).split(',')
        for word_params in related_word_forms_list:
            word_name = word_params.split()[0]
            yield {
                'name': word_name,
                'base_word_name': row[1].value,
                'base_word_frequency': int(row[0].value.replace('k', '')),
            }


def serialize_words_to_pickle(word_storage: Words):
    """Сериализация слов из word_storage в pickle"""
    logger.debug("Начало сериализации слов в pickle формат")
    with open(path_to_serialization, 'wb') as f:
        pickle.dump(word_storage.get_words(), f)
    logger.debug("Сериализация слов в pickle формат завершена")
